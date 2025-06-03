import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import requests
from openpyxl import load_workbook
from datetime import datetime, timedelta

# Function to fetch XLSX file from Artifactory (same as previous)
def fetch_xlsx_from_artifactory(sw_version):
    artifactory_url = 'https://artifactory.yourdomain.com/artifactory'
    repo = 'your-repo'
    file_path = f'{repo}/software_versions/{sw_version}.xlsx'
    
    headers = {
        'Authorization': 'Bearer your_api_key'  # Or use basic auth if needed
    }

    response = requests.get(f'{artifactory_url}/{file_path}', headers=headers)

    if response.status_code == 200:
        with open(f'{sw_version}.xlsx', 'wb') as f:
            f.write(response.content)
        print(f'File {sw_version}.xlsx downloaded successfully.')
        return f'{sw_version}.xlsx'
    else:
        print(f'Error fetching file: {response.status_code}')
        return None

# Function to read the XLSX file and get the data based on selected SW version
def get_approval_data(sw_version_file):
    try:
        wb = load_workbook(sw_version_file)
        sheet = wb.active  # Assuming the data is in the active sheet
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            request_id, requester, status, comments = row
            data.append({
                'request_id': request_id,
                'requester': requester,
                'status': status,
                'comments': comments
            })
        return data
    except Exception as e:
        print(f"Error reading the XLSX file: {e}")
        return []

# Function to send email (same as previous)
def send_email(subject, body, recipient_email):
    sender_email = "your-email@domain.com"
    password = "your-email-password"  # Use app-specific password or OAuth for security

    # Email sending logic...
    # (Same as in previous response)

# Function to send notifications and reminders (same as previous)
def send_notifications_and_reminders(sw_version):
    sw_version_file = fetch_xlsx_from_artifactory(sw_version)
    
    if sw_version_file:
        approval_data = get_approval_data(sw_version_file)

        for entry in approval_data:
            requester = entry['requester']
            status = entry['status']
            comments = entry['comments']

            if status == 'pending':
                subject = f"Approval Request: {entry['request_id']}"
                body = f"Dear {requester},\n\nPlease provide your decision for request ID {entry['request_id']}.\n\nComments: {comments}\n\nKindly respond at your earliest convenience."
                send_email(subject, body, requester)

            reminder_time = datetime.now() - timedelta(days=2)
            if status == 'pending' and datetime.now() > reminder_time:
                reminder_subject = f"Reminder: Approval Request {entry['request_id']} - Decision Pending"
                reminder_body = f"Dear {requester},\n\nThis is a reminder to provide your decision for request ID {entry['request_id']}.\n\nComments: {comments}\n\nPlease respond as soon as possible."
                send_email(reminder_subject, reminder_body, requester)

# GUI Design using Tkinter
class SoftwareApprovalApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Software Approval Management")
        self.root.geometry("600x400")

        # Software version dropdown
        self.sw_version_label = tk.Label(root, text="Select Software Version:")
        self.sw_version_label.pack(pady=10)

        self.sw_version_combobox = ttk.Combobox(root, values=["v1.0", "v2.0", "v3.0"], state="readonly")
        self.sw_version_combobox.pack(pady=10)
        self.sw_version_combobox.current(0)

        # Search entry
        self.search_label = tk.Label(root, text="Search by Requester:")
        self.search_label.pack(pady=10)

        self.search_entry = tk.Entry(root)
        self.search_entry.pack(pady=10)

        # Search button
        self.search_button = tk.Button(root, text="Search Requests", command=self.search_requests)
        self.search_button.pack(pady=10)

        # Display area
        self.results_text = tk.Text(root, height=10, width=70)
        self.results_text.pack(pady=10)

        # Notify and Reminder button
        self.notify_button = tk.Button(root, text="Send Notifications & Reminders", command=self.notify_and_remind)
        self.notify_button.pack(pady=10)

    def search_requests(self):
        sw_version = self.sw_version_combobox.get()
        search_term = self.search_entry.get()

        # Fetch and filter requests based on search term
        sw_version_file = fetch_xlsx_from_artifactory(sw_version)
        if sw_version_file:
            approval_data = get_approval_data(sw_version_file)
            filtered_data = [entry for entry in approval_data if search_term.lower() in entry['requester'].lower()]
            
            # Display the results in the text area
            self.results_text.delete(1.0, tk.END)  # Clear previous results
            for entry in filtered_data:
                self.results_text.insert(tk.END, f"Request ID: {entry['request_id']}\nRequester: {entry['requester']}\nStatus: {entry['status']}\nComments: {entry['comments']}\n\n")

    def notify_and_remind(self):
        sw_version = self.sw_version_combobox.get()
        send_notifications_and_reminders(sw_version)
        messagebox.showinfo("Notification", "Notifications and reminders have been sent.")

# Run the app
if __name__ == "__main__":
    root = tk.Tk()
    app = SoftwareApprovalApp(root)
    root.mainloop()
